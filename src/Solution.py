from typing import List, Dict, Optional
import os
import pathlib
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from RouteSolution import RouteSolution

class Solution:
    """
    输出Excel文件格式的结果。
    Solution类用于存储一组路线,这些路线共同组成完整的时刻表。
    同时提供了有用的I/O功能。
    """
    
    debug_mode: bool = False
    
    def __init__(self, debug_mode: bool):
        """
        构造函数
        Args:
            debug_mode: 是否输出调试信息
        """
        self.route_lists: List[RouteSolution] = []  #包含所有求解后的运行线数据（各峰期 上下行 所有数据）一根线等于一个RouteSolution
        self.debug_mode = debug_mode
        
    def addTrainService(self, rs: RouteSolution) -> None:
        """添加一条运行线路，代表一个实际执行的车次"""
        self.route_lists.append(rs)
        
    def modify_route(self, i: int, rs: RouteSolution) -> None:
        """修改指定索引的路线"""
        self.route_lists[i] = rs
        
    def renumb_routes(self) -> None:
        """
        重新编号路线。
        由于路线编号可能混乱,使用此函数清理编号。
        """
        route_num = 0
        route_map: Dict[int, int] = {}
        
        for rs in self.route_lists:
            rn = rs.car_info.table_num
            # 检查是否已为该列车分配新编号
            if rn not in route_map:
                route_num += 1
                route_map[rn] = route_num
            # 应用新编号    
            rn = route_map[rn]
            rs.car_info.table_num = rn
            rs.car_info.round_num = rs.car_info.round_num % 100 + rn * 100
            # 修改车次号生成逻辑，使用3位数车次号
            # rs.car_info.round_num = (rn % 999) + 1

    def renumb_routes_new(self) -> None:
        """
        重新编号路线。
        由于路线编号可能混乱,使用此函数清理编号。
        """
        route_num = 0
        route_map: Dict[int, int] = {}
        round_counter = 1  # 车次号计数器
        
        for rs in self.route_lists:
            rn = rs.car_info.table_num
            # 检查是否已为该列车分配新编号
            if rn not in route_map:
                route_num += 1
                route_map[rn] = route_num
            # 应用新的表号（车辆编号）    
            rs.car_info.table_num = route_map[rn]
            
            # 为每个车次分配唯一的车次号
            rs.car_info.round_num = round_counter
            round_counter += 1
            
            # 确保车次号不超过999
            if round_counter > 999:
                round_counter = 1       
                 
    def writeCSV(self, file_name: str, rl: 'RailInfo', csn_default: str) -> None:
        """
        将当前解决方案输出到文件
        Args:
            file_name: 输出文件目录
            rl: 包含必要信息的RailInfo对象（即轨道信息数据xml文件）
            csn_default: 默认字符编码
        """
        try:
            pathlib.Path(file_name).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(e)
            
        csn = "gbk"
        if csn_default:
            print(f"Writing files with {csn_default}")
            csn = csn_default
        else:
            os_name = os.name
            if os_name == "posix":
                csn = "utf-8"
                print(f"Writing files under OS: {os_name}, using UTF-8")
                
        # 写入速度数据
        try:
            with open(f"{file_name}/result_spd.csv", "w", encoding=csn) as out:
                header = "开始站台目的地码 结束站台目的地码 第1等级的运行时长 第2等级的运行时长 第3等级的运行时长 第4等级的运行时长 第5等级的运行时长"
                out.write(header + "\n")
                
                hmp: Dict[str, List[int]] = {}
                for tl, val in rl.travel_time_map.items():
                    splited = tl.split("_")
                    spd_level = int(splited[2]) - 1
                    tl = f"{splited[0]}_{splited[1]}"
                    
                    if tl not in hmp:
                        hmp[tl] = [0] * len(rl.speedLevels_name)
                    hmp[tl][spd_level] = val
                    
                for tl, vals in hmp.items():
                    splited = tl.split("_")
                    if self.debug_mode:
                        print(f"{splited[0]}  {splited[1]}")
                    line = f"{splited[0]} {splited[1]} " + " ".join(str(v) for v in vals)
                    out.write(line + "\n")
                    
        except Exception as e:
            print("General error occurred")
            print(e)
            
        # 写入计划数据
        try:
            tmp_map: Dict[int, List[str]] = {}
            with open(f"{file_name}/result_planned.csv", "w", encoding=csn) as out:
                header = "表号 车次号 路径编号 发车时间 自动车次号 快车 载客 列车编号"
                out.write(header + "\n")
                
                for rs in self.route_lists:
                    route_num = rs.car_info.route_num
                    if route_num not in tmp_map:
                        tmp_map[route_num] = []
                    tmp_map[route_num].append(rs.retCSVStringPlanned())
                    
                for tmp_list in tmp_map.values():
                    for line in tmp_list:
                        out.write(line + "\n")
                        
        except Exception as e:
            print("General error occurred")
            print(e)
            
        # 写入任务数据
        try:
            res_map: Dict[int, str] = {}
            with open(f"{file_name}/result_mission.csv", "w", encoding=csn) as out:
                header = "表号 车次号 站台目的地码 到站时间 离站时间 运行等级 是否清客"
                out.write(header + "\n")
                
                for rs in self.route_lists:
                    tmp_map = rs.retCSVStringMission()
                    res_map.update(tmp_map)
                    
                for x in sorted(res_map.keys()):
                    out.write(res_map[x])
                    
        except Exception as e:
            print("General error occurred")
            print(e)
            
    def writeCSV_num(self, file_name: str, rl: 'RailInfo', csn_default: str) -> None:
        """与writeCSV相同,但不格式化日期"""
        # 实现与writeCSV相同,但调用retCSVStringPlanned_num和retCSVStringMission_num
        self.writeCSV(file_name, rl, csn_default)
        
    def generateSpeedSheet(self, rl: 'RailInfo', sheet: Worksheet) -> None:
        """生成速度表格"""
        # 表头
        headers = [
            "开始站台目的地码", "结束站台目的地码",
            "第1等级的运行时长", "第2等级的运行时长", "第3等级的运行时长",
            "第4等级的运行时长", "第5等级的运行时长"
        ]
        for i, header in enumerate(headers):
            sheet.cell(row=1, column=i+1, value=header)
            
        # 获取速度数据
        # 创建一个字典 hmp 用于存储速度数据，其中：
        # - 键(key)是由"起始站台目的地码_终点站台目的地码"组成的字符串
        # - 值(value)是一个列表，包含了不同速度等级下的运行时长
        # 数据来源：从 `'RailInfo'.py` 中的 travel_time_map 获取原始数据，这个 map 存储了每一段路径在不同速度等级下的运行时间。
        hmp: Dict[str, List[int]] = {}
        # 数据处理过程：
        # - 遍历 travel_time_map 中的每一条记录
        # - 将键(形如"起点_终点_速度等级")拆分，提取出起点和终点信息
        # - 根据速度等级(spd_level)，将运行时间(val)存储到对应的位置
        for tl, val in rl.travel_time_map.items():
            splited = tl.split("_")
            spd_level = int(splited[2]) - 1
            tl = f"{splited[0]}_{splited[1]}"
            
            if tl not in hmp:
                hmp[tl] = [0] * len(rl.speedLevels_name)
            hmp[tl][spd_level] = val
            
        # 逐行设置单元格值
        # 最终这些数据会被用于生成 Excel 表格中的"运行时间"工作表，包含了各个站点之间在不同速度等级下的运行时间数据。
        row_num = 2#从第2行开始写入数据（ row_num = 2 ，因为第1行是表头）
        for tl, vals in hmp.items():#遍历之前处理好的速度数据字典 hmp
            # - tl 是键，形如"起点站台码_终点站台码"
            # - vals 是值，是一个列表，包含了该站点对之间在不同速度等级下的运行时长
            splited = tl.split("_")#将站点对信息( tl )按下划线分割，得到起点和终点站台码
            if self.debug_mode:#如果开启了调试模式，会打印起点和终点站台码
                print(f"{splited[0]}  {splited[1]}")
            #在Excel表格中：- 第1列写入起点站台码- 第2列写入终点站台码    
            sheet.cell(row=row_num, column=1, value=splited[0])
            sheet.cell(row=row_num, column=2, value=splited[1])
            # 从第3列开始，依次写入不同速度等级下的运行时长
            for i, val in enumerate(vals):
                sheet.cell(row=row_num, column=i+3, value=val)
            # 每处理完一行数据，行号加1，继续处理下一组数据    
            row_num += 1
            
    def generateRouteSheet(self, rl: 'RailInfo', sheet: Worksheet) -> Dict[int, List[int]]:
        """生成计划线数据表格"""
        if not self.route_lists:
            print("警告：route_lists为空，没有可用的路线数据")
        tmp_map: Dict[int, Dict[int, str]] = {}# 存储每个车次的发车时间和对应的计划数据
        indx_map: Dict[int, Dict[int, int]] = {}# 存储每个车次的发车时间和对应的索引
        indx_map2: Dict[int, List[int]] = {}# 存储每个车次的索引列表
        # 添加打印语句，查看route_lists的内容
        # print(f"route_lists的内容: {self.route_lists}")
        print(f"route_lists长度: {len(self.route_lists)}")

        # 表头（计划线数据）
        headers = ["表号", "车次号", "路径编号", "发车时间", 
                  "自动车次号", "快车", "载客", "列车编号"]
        for i, header in enumerate(headers):
            sheet.cell(row=1, column=i+1, value=header)
            # print("填充表头")
        # 获取数据
        for i, rs in enumerate(self.route_lists):
            try:
                route_num = rs.car_info.table_num
                if route_num not in tmp_map:
                    tmp_map[route_num] = {}
                    indx_map[route_num] = {}
                    indx_map2[route_num] = []
                
                str_data = rs.retCSVStringPlanned_num()
                splited = str_data.split()
                # 确保发车时间是有效的整数字符串
                send_time = int(float(splited[3].strip('[].')))  # 处理可能的方括号和小数点
                
                if send_time not in tmp_map[route_num]:
                    tmp_map[route_num][send_time] = str_data
                if send_time not in indx_map[route_num]:
                    indx_map[route_num][send_time] = i
                    
                # print(f"处理第{i}条运行线：route_num={route_num}, send_time={send_time}")
                
            except Exception as e:
                print(f"处理第{i}条运行线时出错：{e}")
                print(f"错误数据：{str_data if 'str_data' in locals() else 'N/A'}")
                continue
        # 生成行 主要负责将处理好的数据写入Excel表格
        # print("填充计划线数据：准备中2")
        row_num = 2#从第2行开始写入数据（因为第1行是表头）
        for route_num, tmp_list in tmp_map.items():
            # print("填充计划线数据0")
            for stime in sorted(tmp_list.keys()):#按照表号（route_num）和发车时间（stime）的顺序组织数据
                # 对于每个表号下的数据：
                # - 按发车时间排序（使用 sorted(tmp_list.keys()) ）
                # - 将每条计划数据分割成单独的字段
                # - 将前8个字段写入对应的单元格（对应表头中的8个列）
                # print("填充计划线数据1")
                splited = tmp_list[stime].split()
                for j, val in enumerate(splited[:8]):
                    sheet.cell(row=row_num, column=j+1, value=val)
                    # print("填充计划线数据2")
                row_num += 1
                # - 同时维护了一个索引映射表 indx_map2
                # - 对于每个表号，按照发车时间顺序记录其在原始数据（route_lists）中的索引位置
                # - 这个索引映射表会被返回并用于后续生成任务数据表格
                indx_map2[route_num].append(indx_map[route_num][stime])
                
        return indx_map2
        
    def generateMissionSheet(self, rl: 'RailInfo', sheet: Worksheet, 
                           index_map: Dict[int, List[int]]) -> None:
        """生成任务表格"""
        # 表头
        headers = ["表号", "车次号", "站台目的地码", "到站时间", 
                  "离站时间", "运行等级", "是否清客"]
        for i, header in enumerate(headers):#将这些标题写入Excel表格的第一行
            sheet.cell(row=1, column=i+1, value=header)
        # 获取数据
        row_num = 2#从第2行开始写入数据（因为第1行是表头）
        for t, ind_list in index_map.items():#使用从 generateRouteSheet 返回的 index_map 作为数据索引
            #index_map 是一个字典，键是表号，值是该表号下所有列车在 route_lists 中的索引列表
            # print("填充任务线数据1")   
            for idx in ind_list:#按表号遍历 index_map
                # print("填充任务线数据2")   
                # 对每个表号下的索引列表：
                # - 使用索引从 route_lists 获取对应的 RouteSolution 对象
                # 在 `RouteSolution.py` 中， retCSVStringMission 方法需要以下数据才能正常工作：
                # - stopped_platforms (停靠的站台码)
                # - arr_time (到达时间)
                # - dep_time (出发时间)
                # - performance_levels (运行等级)
                # - car_info (车辆信息)
                # - 调用 retCSVStringMission_num() 获取该列车的任务数据
                # - 将获取的数据按时间顺序排序（使用 sorted(tmp_map.keys()) ）
                # - 将每条任务数据分割成单独的字段
                # - 将前7个字段写入对应的单元格
                tmp_map = self.route_lists[idx].retCSVStringMission_num()
                # print(f"tmp_map: {tmp_map}")
                #如果tmp_map为空，这个循环进不去，无法填充计划线数据
                for x in sorted(tmp_map.keys()):
                    # print("填充任务线数据3")
                    tmp = tmp_map[x]
                    splited = tmp.split()
                    for j, val in enumerate(splited[:7]):
                        # print("填充任务线数据4")
                        sheet.cell(row=row_num, column=j+1, value=val)
                    row_num += 1
                    
    def writeExcel(self, file_name: str, rl: 'RailInfo', csn_default: str) -> None:
        """输出Excel格式结果"""
        try:
            pathlib.Path(file_name).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(e)
            
        csn = "gbk"
        if csn_default:
            print(f"Writing files with {csn_default}")
            csn = csn_default
        else:
            os_name = os.name
            if os_name == "posix":
                csn = "utf-8"
                print(f"Writing files under OS: {os_name}, using UTF-8")
                
        try:
            workbook = Workbook()
            
            # 创建速度表
            sheet = workbook.active
            sheet.title = "运行时间"
            self.generateSpeedSheet(rl, sheet)
            print("创建运行等级速度表完成")
            # 创建计划线数据表
            sheet1 = workbook.create_sheet("计划线数据")
            index_map = self.generateRouteSheet(rl, sheet1)
            print("创建计划线数据表完成")
            # 创建任务线数据表
            sheet2 = workbook.create_sheet("任务线数据")
            self.generateMissionSheet(rl, sheet2, index_map)
            print("创建任务线数据表完成")
            # 保存文件
            workbook.save(f"{file_name}/result.xls")
            
        except Exception as e:
            print(e)
            import sys
            sys.exit(0)