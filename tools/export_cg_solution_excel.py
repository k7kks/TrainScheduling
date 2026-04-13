import argparse
import csv
import json
import shutil
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = REPO_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

try:
    from Engineering import Engineering  # type: ignore
    from CarInfo import CarInfo  # type: ignore
    from RouteSolution import RouteSolution  # type: ignore
    from Platform import PlatformType  # type: ignore
except Exception:
    Engineering = None  # type: ignore
    CarInfo = None  # type: ignore
    RouteSolution = None  # type: ignore
    PlatformType = None  # type: ignore


DEPOT_ROUTE_TIME_EARLY = 28 * 3600


def _read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path: Path, headers: Sequence[str], rows: Iterable[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(headers))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_sheet(sheet, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


def _parse_int(value: object, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return default
    return int(float(text))


def _compute_send_time_in_connect(engine, route_solution) -> int:
    if route_solution is None or not getattr(route_solution, "dep_time", None):
        return 0
    base_time = _parse_int(route_solution.dep_time[-1])
    if engine is None or PlatformType is None:
        return base_time
    try:
        stopped_platforms = getattr(route_solution, "stopped_platforms", [])
        if not stopped_platforms:
            return base_time
        final_platform = engine.rl.platformList[stopped_platforms[-1]]
        if getattr(final_platform, "platform_type", None) == PlatformType.TURNBACK:
            turnback = engine.rl.turnbackList[final_platform.dest_code]
            base_time += _parse_int(getattr(turnback, "min_tb_time", 0))
    except Exception:
        return base_time
    return base_time


def _load_summary(solution_dir: Path) -> Dict[str, object]:
    summary_path = solution_dir / "summary.json"
    if not summary_path.exists():
        return {}
    for encoding in ("utf-8", "utf-8-sig", "gbk", "gb18030"):
        try:
            return json.loads(summary_path.read_text(encoding=encoding))
        except Exception:
            continue
    return {}


def _load_manifest(instance_dir: Path) -> Dict[str, str]:
    manifest_path = instance_dir / "manifest.txt"
    if not manifest_path.exists():
        return {}
    manifest: Dict[str, str] = {}
    for raw_line in manifest_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or "=" not in line:
            continue
        key, value = line.split("=", 1)
        manifest[key.strip()] = value.strip()
    return manifest


def _build_export_engine(instance_dir: Path):
    if Engineering is None:
        return None
    manifest = _load_manifest(instance_dir)
    rail_path = manifest.get("rail_path", "")
    setting_path = manifest.get("setting_path", "")
    if not rail_path or not setting_path:
        return None
    allow_mixed = _parse_int(manifest.get("allow_mixed", 0))
    try:
        eng = Engineering(
            False,
            str(Path(rail_path)),
            str(Path(setting_path)),
            algorithm_type="LP",
            solver_type="PULP",
            obj_type=1,
            arrange_init=False,
            n_phase=1,
            allow_mixed_operation=allow_mixed,
        )
        eng.allow_mixed_operation = allow_mixed
        eng.phase0()
        return eng
    except Exception:
        return None


def _group_slot_missions(slot_mission_rows: List[Dict[str, str]]) -> Dict[int, List[Dict[str, str]]]:
    missions_by_slot: Dict[int, List[Dict[str, str]]] = {}
    for row in slot_mission_rows:
        slot_id = _parse_int(row["slot_id"])
        missions_by_slot.setdefault(slot_id, []).append(row)
    for rows in missions_by_slot.values():
        rows.sort(key=lambda item: _parse_int(item["stop_index"]))
    return missions_by_slot


def _sorted_selected_trips(selected_trip_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        selected_trip_rows,
        key=lambda row: (
            _parse_int(row["vehicle_id"]),
            _parse_int(row["sequence_index"]),
            _parse_int(row["selected_departure"]),
        ),
    )


def _build_mainline_route_solution(trip_row: Dict[str, str], missions_by_slot: Dict[int, List[Dict[str, str]]]):
    if RouteSolution is None or CarInfo is None:
        return None
    slot_id = _parse_int(trip_row["slot_id"])
    shift_seconds = _parse_int(trip_row.get("shift_seconds", 0))
    rs = RouteSolution(_parse_int(trip_row["selected_departure"]))
    rs.car_info = CarInfo(
        _parse_int(trip_row["vehicle_id"]),
        _parse_int(trip_row["sequence_index"]) + 1,
        _parse_int(trip_row["route_id"]),
    )
    rs.dir = _parse_int(trip_row["direction"])
    rs.xroad = _parse_int(trip_row["xroad"])
    rs.phase = _parse_int(trip_row["phase_id"])
    for mission_row in missions_by_slot.get(slot_id, []):
        arrival = _parse_int(mission_row["arrival"]) + shift_seconds
        departure = _parse_int(mission_row["departure"]) + shift_seconds
        rs.addStop(
            str(mission_row["platform_code"]),
            max(0, departure - arrival),
            _parse_int(mission_row.get("perf_level", 1), 1),
            arrival,
            departure,
        )
    if rs.car_info is not None:
        rs.car_info.arr_time = _parse_int(trip_row["selected_arrival"])
    return rs


def _decorate_generated_route(route_solution, template_route, route_id: str) -> None:
    route_solution.car_info.route_num = int(route_id)
    route_solution.car_info.table_num = int(template_route.car_info.table_num)
    route_solution.dir = int(template_route.dir)
    route_solution.xroad = int(template_route.xroad)
    route_solution.phase = int(template_route.phase)
    route_solution.operating = False
    route_solution.car_info.arr_time = int(route_solution.dep_time[-1]) if route_solution.dep_time else 0


def _build_outbound_depot_routes(eng, first_route) -> List[object]:
    if (
        eng is None
        or CarInfo is None
        or first_route is None
        or first_route.car_info is None
    ):
        return []
    xroad = int(first_route.xroad)
    direction = int(first_route.dir)
    depot_infos = getattr(eng.us, "depot_routes_infos", [])
    if xroad < 0 or xroad >= len(depot_infos):
        return []
    depot_info = depot_infos[xroad]
    if len(depot_info.routes[1]) <= direction:
        return []
    route_chain = [str(route_id) for route_id in depot_info.routes[1][direction] if str(route_id).strip()]
    if not route_chain:
        return []

    speed_level = int(getattr(eng, "level", -1))
    next_departure = int(first_route.dep_time[0])
    built_routes_reversed: List[object] = []
    for route_id in reversed(route_chain):
        probe_rs = eng.rl.generateSinglePathSolution(
            route_id,
            0,
            speed_level,
            0,
            DEPOT_ROUTE_TIME_EARLY,
            CarInfo(first_route.car_info.table_num, first_route.car_info.round_num, int(route_id)),
        )
        if probe_rs is None or not probe_rs.dep_time:
            continue
        duration = int(probe_rs.dep_time[-1]) - int(probe_rs.dep_time[0])
        send_time = next_departure - duration
        depot_rs = eng.rl.generateSinglePathSolution(
            route_id,
            send_time,
            speed_level,
            0,
            DEPOT_ROUTE_TIME_EARLY,
            CarInfo(first_route.car_info.table_num, first_route.car_info.round_num, int(route_id)),
        )
        if depot_rs is None:
            continue
        _decorate_generated_route(depot_rs, first_route, route_id)
        built_routes_reversed.append(depot_rs)
        next_departure = send_time
    built_routes_reversed.reverse()
    return built_routes_reversed


def _build_inbound_depot_routes(eng, last_route) -> List[object]:
    if (
        eng is None
        or CarInfo is None
        or last_route is None
        or last_route.car_info is None
    ):
        return []
    xroad = int(last_route.xroad)
    direction = int(last_route.dir)
    depot_infos = getattr(eng.us, "depot_routes_infos", [])
    if xroad < 0 or xroad >= len(depot_infos):
        return []
    depot_info = depot_infos[xroad]
    if len(depot_info.routes[0]) <= direction:
        return []
    route_chain = [str(route_id) for route_id in depot_info.routes[0][direction] if str(route_id).strip()]
    if not route_chain:
        return []

    speed_level = int(getattr(eng, "level", -1))
    built_routes: List[object] = []
    previous_route = last_route
    for route_id in route_chain:
        send_time = _compute_send_time_in_connect(eng, previous_route)
        depot_rs = eng.rl.generateSinglePathSolution(
            route_id,
            send_time,
            speed_level,
            0,
            DEPOT_ROUTE_TIME_EARLY,
            CarInfo(last_route.car_info.table_num, last_route.car_info.round_num, int(route_id)),
        )
        if depot_rs is None:
            continue
        _decorate_generated_route(depot_rs, last_route, route_id)
        built_routes.append(depot_rs)
        previous_route = depot_rs
    return built_routes


def _assign_duty_export_numbers(vehicle_id: int, duty_routes: List[object]) -> None:
    if CarInfo is None:
        return
    total = len(duty_routes)
    for index, route_solution in enumerate(duty_routes):
        round_num = vehicle_id * 100 + (total - index)
        if route_solution.car_info is None:
            route_solution.car_info = CarInfo(vehicle_id, round_num, 0)
        route_solution.car_info.table_num = vehicle_id
        route_solution.car_info.round_num = round_num


def _route_solution_to_planned_row(route_solution) -> Dict[str, object]:
    return {
        "琛ㄥ彿": _parse_int(route_solution.car_info.table_num),
        "杞︽鍙?": _parse_int(route_solution.car_info.round_num),
        "璺緞缂栧彿": _parse_int(route_solution.car_info.route_num),
        "鍙戣溅鏃堕棿": _parse_int(route_solution.dep_time[0]),
        "鑷姩杞︽鍙?": 1,
        "蹇溅": 0,
        "杞藉": 0,
        "鍒楄溅缂栧彿": 1,
    }


def _route_solution_to_mission_rows(route_solution) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    stop_count = len(route_solution.stopped_platforms)
    for index, platform_code in enumerate(route_solution.stopped_platforms):
        arrival = _parse_int(route_solution.dep_time[0]) if index == 0 else _parse_int(route_solution.arr_time[index])
        departure = (
            _parse_int(route_solution.dep_time[index])
            if index < stop_count - 1
            else _parse_int(route_solution.arr_time[index])
        )
        rows.append(
            {
                "琛ㄥ彿": _parse_int(route_solution.car_info.table_num),
                "杞︽鍙?": _parse_int(route_solution.car_info.round_num),
                "绔欏彴鐩殑鍦扮爜": str(platform_code),
                "鍒扮珯鏃堕棿": arrival,
                "绂荤珯鏃堕棿": departure,
                "杩愯绛夌骇": _parse_int(route_solution.performance_levels[index]),
                "鏄惁娓呭": 1 if index == stop_count - 1 else 0,
            }
        )
    return rows


def _build_legacy_like_rows(
    instance_dir: Path,
    selected_trip_rows: List[Dict[str, str]],
    slot_mission_rows: List[Dict[str, str]],
) -> Optional[Tuple[List[str], List[Dict[str, object]], List[List[object]], List[str], List[Dict[str, object]], List[List[object]]]]:
    eng = _build_export_engine(instance_dir)
    if eng is None or RouteSolution is None or CarInfo is None:
        return None

    planned_headers = ["琛ㄥ彿", "杞︽鍙?", "璺緞缂栧彿", "鍙戣溅鏃堕棿", "鑷姩杞︽鍙?", "蹇溅", "杞藉", "鍒楄溅缂栧彿"]
    mission_headers = ["琛ㄥ彿", "杞︽鍙?", "绔欏彴鐩殑鍦扮爜", "鍒扮珯鏃堕棿", "绂荤珯鏃堕棿", "杩愯绛夌骇", "鏄惁娓呭"]

    missions_by_slot = _group_slot_missions(slot_mission_rows)
    duties: Dict[Tuple[int, int], List[Dict[str, str]]] = {}
    for trip_row in _sorted_selected_trips(selected_trip_rows):
        key = (_parse_int(trip_row["vehicle_id"]), _parse_int(trip_row["column_id"]))
        duties.setdefault(key, []).append(trip_row)

    planned_csv_rows: List[Dict[str, object]] = []
    planned_sheet_rows: List[List[object]] = []
    mission_csv_rows: List[Dict[str, object]] = []
    mission_sheet_rows: List[List[object]] = []

    for (vehicle_id, _column_id), trip_rows in sorted(duties.items()):
        mainline_routes = []
        for trip_row in trip_rows:
            mainline_route = _build_mainline_route_solution(trip_row, missions_by_slot)
            if mainline_route is not None:
                mainline_routes.append(mainline_route)
        if not mainline_routes:
            continue

        duty_routes: List[object] = []
        duty_routes.extend(_build_outbound_depot_routes(eng, mainline_routes[0]))
        duty_routes.extend(mainline_routes)
        duty_routes.extend(_build_inbound_depot_routes(eng, mainline_routes[-1]))

        for index in range(len(duty_routes) - 1):
            duty_routes[index].next_ptr = duty_routes[index + 1]
            duty_routes[index + 1].prev_ptr = duty_routes[index]

        _assign_duty_export_numbers(vehicle_id, duty_routes)
        for route_solution in duty_routes:
            planned_row = _route_solution_to_planned_row(route_solution)
            planned_csv_rows.append(planned_row)
            planned_sheet_rows.append([planned_row[key] for key in planned_headers])
            for mission_row in _route_solution_to_mission_rows(route_solution):
                mission_csv_rows.append(mission_row)
                mission_sheet_rows.append([mission_row[key] for key in mission_headers])

    return (
        planned_headers,
        planned_csv_rows,
        planned_sheet_rows,
        mission_headers,
        mission_csv_rows,
        mission_sheet_rows,
    )


def _build_speed_rows(travel_time_rows: List[Dict[str, str]]) -> Tuple[List[str], List[List[object]]]:
    level_headers = sorted([key for key in travel_time_rows[0].keys() if key.startswith("level_")], key=lambda item: int(item.split("_")[1])) if travel_time_rows else []
    headers = [
        "开始站台目的地码",
        "结束站台目的地码",
    ] + [f"第{idx}等级的运行时长" for idx in range(1, len(level_headers) + 1)]
    rows: List[List[object]] = []
    for row in sorted(travel_time_rows, key=lambda item: (item["start_platform"], item["end_platform"])):
        rows.append([row["start_platform"], row["end_platform"]] + [_parse_int(row.get(level_name, 0)) for level_name in level_headers])
    return headers, rows


def _build_planned_rows(selected_trip_rows: List[Dict[str, str]]) -> Tuple[List[str], List[Dict[str, object]], List[List[object]]]:
    headers = ["表号", "车次号", "路径编号", "发车时间", "自动车次号", "快车", "载客", "列车编号"]
    ordered_rows = sorted(
        selected_trip_rows,
        key=lambda row: (
            _parse_int(row["vehicle_id"]),
            _parse_int(row["selected_departure"]),
            _parse_int(row["sequence_index"]),
        ),
    )
    csv_rows: List[Dict[str, object]] = []
    sheet_rows: List[List[object]] = []
    for row in ordered_rows:
        planned_row = {
            "表号": _parse_int(row["vehicle_id"]),
            "车次号": _parse_int(row["round_num"]),
            "路径编号": _parse_int(row["route_id"]),
            "发车时间": _parse_int(row["selected_departure"]),
            "自动车次号": 1,
            "快车": 0,
            "载客": 0,
            "列车编号": 1,
        }
        csv_rows.append(planned_row)
        sheet_rows.append([planned_row[key] for key in headers])
    return headers, csv_rows, sheet_rows


def _build_mission_rows(
    selected_trip_rows: List[Dict[str, str]],
    slot_mission_rows: List[Dict[str, str]],
) -> Tuple[List[str], List[Dict[str, object]], List[List[object]]]:
    headers = ["表号", "车次号", "站台目的地码", "到站时间", "离站时间", "运行等级", "是否清客"]
    missions_by_slot: Dict[int, List[Dict[str, str]]] = {}
    for row in slot_mission_rows:
        slot_id = _parse_int(row["slot_id"])
        missions_by_slot.setdefault(slot_id, []).append(row)
    for rows in missions_by_slot.values():
        rows.sort(key=lambda item: _parse_int(item["stop_index"]))

    ordered_rows = sorted(
        selected_trip_rows,
        key=lambda row: (
            _parse_int(row["vehicle_id"]),
            _parse_int(row["selected_departure"]),
            _parse_int(row["sequence_index"]),
        ),
    )
    csv_rows: List[Dict[str, object]] = []
    sheet_rows: List[List[object]] = []
    for trip_row in ordered_rows:
        slot_id = _parse_int(trip_row["slot_id"])
        shift_seconds = _parse_int(trip_row.get("shift_seconds", 0))
        vehicle_id = _parse_int(trip_row["vehicle_id"])
        round_num = _parse_int(trip_row["round_num"])
        for mission_row in missions_by_slot.get(slot_id, []):
            row = {
                "表号": vehicle_id,
                "车次号": round_num,
                "站台目的地码": mission_row["platform_code"],
                "到站时间": _parse_int(mission_row["arrival"]) + shift_seconds,
                "离站时间": _parse_int(mission_row["departure"]) + shift_seconds,
                "运行等级": _parse_int(mission_row["perf_level"]),
                "是否清客": _parse_int(mission_row["clear_flag"]),
            }
            csv_rows.append(row)
            sheet_rows.append([row[key] for key in headers])
    return headers, csv_rows, sheet_rows


def _load_peaks(instance_dir: Path) -> List[Dict[str, int]]:
    peaks = []
    for row in _read_csv(instance_dir / "peaks.csv"):
        peaks.append(
            {
                "peak_id": _parse_int(row["peak_id"]),
                "start_time": _parse_int(row["start_time"]),
                "end_time": _parse_int(row["end_time"]),
                "train_num": _parse_int(row["train_num"]),
            }
        )
    return peaks


def _peak_incidence_from_rows(rows: Iterable[Dict[str, object]], departure_key: str, vehicle_key: str, peaks: List[Dict[str, int]]) -> Dict[int, int]:
    peak_to_vehicles: Dict[int, set] = {peak["peak_id"]: set() for peak in peaks}
    for row in rows:
        departure = _parse_int(row[departure_key])
        vehicle = str(row[vehicle_key])
        for peak in peaks:
            if peak["start_time"] <= departure < peak["end_time"]:
                peak_to_vehicles[peak["peak_id"]].add(vehicle)
    return {peak_id: len(vehicles) for peak_id, vehicles in peak_to_vehicles.items()}


def _load_baseline_metrics(baseline_xlsx: Path, peaks: List[Dict[str, int]]) -> Dict[str, object]:
    workbook = load_workbook(baseline_xlsx, read_only=True, data_only=True)
    planned_rows: List[Dict[str, object]] = []
    mission_rows = 0
    if "计划线数据" in workbook.sheetnames:
        sheet = workbook["计划线数据"]
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            headers = [str(value) if value is not None else "" for value in rows[0]]
            for raw_row in rows[1:]:
                if raw_row[0] is None:
                    continue
                planned_rows.append({headers[idx]: raw_row[idx] for idx in range(len(headers))})
    if "任务线数据" in workbook.sheetnames:
        sheet = workbook["任务线数据"]
        mission_rows = max(0, sheet.max_row - 1)
    table_count = len({str(row["表号"]) for row in planned_rows if row.get("表号") is not None})
    return {
        "path": str(baseline_xlsx),
        "planned_rows": len(planned_rows),
        "mission_rows": mission_rows,
        "vehicle_count": table_count,
        "peak_incidence": _peak_incidence_from_rows(planned_rows, "发车时间", "表号", peaks),
    }


def _build_current_metrics(
    selected_trip_rows: List[Dict[str, str]],
    selected_duty_rows: List[Dict[str, str]],
    option_rows: List[Dict[str, str]],
    mission_row_count: int,
    peaks: List[Dict[str, int]],
    summary: Dict[str, object],
    first_car_rows: List[Dict[str, str]],
) -> Dict[str, object]:
    current_peak_incidence = _peak_incidence_from_rows(selected_trip_rows, "selected_departure", "vehicle_id", peaks)
    selected_by_slot = {(_parse_int(row["slot_id"])): row for row in selected_trip_rows}
    options_by_id = {(_parse_int(row["option_id"])): row for row in option_rows}
    first_car_checks = []
    for row in first_car_rows:
        slot_id = _parse_int(row["slot_id"])
        chosen = selected_by_slot.get(slot_id)
        if chosen is None:
            continue
        target_dep = _parse_int(row["target_departure"])
        chosen_option = options_by_id.get(_parse_int(chosen["option_id"]))
        actual_dep = _parse_int(chosen_option.get("first_real_departure") if chosen_option is not None else chosen["selected_departure"])
        first_car_checks.append(
            {
                "slot_id": slot_id,
                "direction": _parse_int(row["direction"]),
                "route_id": row["route_id"],
                "target_platform": row.get("target_platform", ""),
                "target_departure": target_dep,
                "actual_departure": actual_dep,
                "non_early_ok": actual_dep >= target_dep,
            }
        )
    trip_distribution: Dict[int, int] = {}
    for row in selected_duty_rows:
        trip_count = _parse_int(row["trip_count"])
        trip_distribution[trip_count] = trip_distribution.get(trip_count, 0) + 1
    return {
        "planned_rows": len(selected_trip_rows),
        "mission_rows": mission_row_count,
        "vehicle_count": len({_parse_int(row["vehicle_id"]) for row in selected_trip_rows}),
        "max_vehicle_count": summary.get("max_vehicle_count"),
        "vehicle_cap_slack": summary.get("vehicle_cap_slack"),
        "peak_incidence": current_peak_incidence,
        "peak_caps": {peak["peak_id"]: peak["train_num"] for peak in peaks},
        "peak_slack_total": summary.get("peak_slack_total"),
        "peak_slacks": summary.get("peak_slacks", []),
        "headway_target_dev": summary.get("headway_target_dev"),
        "first_car_checks": first_car_checks,
        "trip_distribution": trip_distribution,
    }


def _check_turnback_overlaps(
    solution_dir: Path,
) -> List[Dict[str, object]]:
    """Detect turnback-box overlaps in the selected solution.

    For every turnback platform, reconstruct occupancy windows
    ``[effective_arrival, to_departure]`` and report pairs that overlap
    across different vehicles.
    """
    turnback_file = solution_dir / "selected_turnbacks.csv"
    if not turnback_file.exists():
        return []
    turnback_rows = _read_csv(turnback_file)
    if not turnback_rows:
        return []

    trips_file = solution_dir / "selected_trips.csv"
    if not trips_file.exists():
        return []
    trips_rows = _read_csv(trips_file)

    # option_id → selected_departure
    departure_by_option: Dict[int, int] = {}
    for row in trips_rows:
        departure_by_option[_parse_int(row["option_id"])] = _parse_int(row["selected_departure"])

    # Build turnback occupancy boxes grouped by platform.
    boxes_by_platform: Dict[str, List[Dict[str, object]]] = {}
    for row in turnback_rows:
        platform = row.get("turnback_platform", "")
        if not platform:
            continue
        to_opt_id = _parse_int(row["to_option_id"])
        to_dep = departure_by_option.get(to_opt_id)
        if to_dep is None:
            continue
        wait_time = _parse_int(row.get("wait_time", 0))
        eff_arrival = to_dep - wait_time
        boxes_by_platform.setdefault(platform, []).append(
            {
                "vehicle_id": _parse_int(row["vehicle_id"]),
                "platform": platform,
                "occupy_start": eff_arrival,
                "occupy_end": to_dep,
            }
        )

    overlaps: List[Dict[str, object]] = []
    for platform, boxes in boxes_by_platform.items():
        boxes.sort(key=lambda b: b["occupy_start"])
        for i in range(len(boxes)):
            for j in range(i + 1, len(boxes)):
                if boxes[j]["vehicle_id"] == boxes[i]["vehicle_id"]:
                    continue
                if boxes[j]["occupy_start"] >= boxes[i]["occupy_end"]:
                    break
                overlaps.append(
                    {
                        "platform": platform,
                        "vehicle_a": boxes[i]["vehicle_id"],
                        "vehicle_b": boxes[j]["vehicle_id"],
                        "start_a": boxes[i]["occupy_start"],
                        "end_a": boxes[i]["occupy_end"],
                        "start_b": boxes[j]["occupy_start"],
                        "end_b": boxes[j]["occupy_end"],
                        "overlap_seconds": boxes[i]["occupy_end"] - boxes[j]["occupy_start"],
                    }
                )
    return overlaps


def _write_comparison_report(
    report_path: Path,
    current_metrics: Dict[str, object],
    baseline_metrics: Optional[Dict[str, object]],
    turnback_overlaps: Optional[List[Dict[str, object]]] = None,
) -> None:
    lines = [
        "# CG Solution Comparison",
        "",
        "## Current C++ CG Result",
        "",
        f"- Planned rows: `{current_metrics['planned_rows']}`",
        f"- Mission rows: `{current_metrics['mission_rows']}`",
        f"- Duty count: `{current_metrics['vehicle_count']}`",
        f"- Max vehicle count: `{current_metrics.get('max_vehicle_count')}`",
        f"- Vehicle-cap slack: `{current_metrics.get('vehicle_cap_slack')}`",
        f"- Headway target deviation: `{current_metrics.get('headway_target_dev')}`",
        f"- Peak slack total: `{current_metrics.get('peak_slack_total')}`",
        "",
        "### Peak Incidence",
        "",
        "| Peak | Cap | Current |",
        "| --- | ---: | ---: |",
    ]
    peak_caps = current_metrics.get("peak_caps", {})
    current_peak_incidence = current_metrics.get("peak_incidence", {})
    for peak_id in sorted(peak_caps.keys()):
        lines.append(f"| {peak_id} | {peak_caps[peak_id]} | {current_peak_incidence.get(peak_id, 0)} |")

    first_car_checks = current_metrics.get("first_car_checks", [])
    if first_car_checks:
        lines.extend(
            [
                "",
                "### First-Car Checks",
                "",
                "| Direction | Route | Platform | Target | Actual | Non-early |",
                "| --- | --- | --- | ---: | ---: | --- |",
            ]
        )
        for row in first_car_checks:
            lines.append(
                f"| {row['direction']} | {row['route_id']} | {row['target_platform']} | {row['target_departure']} | {row['actual_departure']} | {row['non_early_ok']} |"
            )

    trip_distribution = current_metrics.get("trip_distribution", {})
    if trip_distribution:
        lines.extend(
            [
                "",
                "### Duty Trip Distribution",
                "",
                "| Trip count | Duty count |",
                "| ---: | ---: |",
            ]
        )
        for trip_count in sorted(trip_distribution.keys()):
            lines.append(f"| {trip_count} | {trip_distribution[trip_count]} |")

    # --- Turnback overlap analysis ---
    if turnback_overlaps is not None:
        lines.extend(
            [
                "",
                "### Turnback Overlap Analysis",
                "",
                f"- Detected overlaps: `{len(turnback_overlaps)}`",
            ]
        )
        if turnback_overlaps:
            lines.extend(
                [
                    "",
                    "| Platform | Vehicle A | Vehicle B | Overlap (s) |",
                    "| --- | ---: | ---: | ---: |",
                ]
            )
            for row in turnback_overlaps[:30]:
                lines.append(
                    f"| {row['platform']} | {row['vehicle_a']} | {row['vehicle_b']} | {row['overlap_seconds']} |"
                )
            if len(turnback_overlaps) > 30:
                lines.append(f"| ... | | | ({len(turnback_overlaps) - 30} more) |")
        else:
            lines.append("- No turnback-box overlaps detected; the solution is conflict-free.")

    if baseline_metrics is not None:
        lines.extend(
            [
                "",
                "## Baseline Workbook",
                "",
                f"- Baseline path: `{baseline_metrics['path']}`",
                f"- Planned rows: `{baseline_metrics['planned_rows']}`",
                f"- Mission rows: `{baseline_metrics['mission_rows']}`",
                f"- Unique table count: `{baseline_metrics['vehicle_count']}`",
                "",
                "### Baseline Peak Incidence",
                "",
                "| Peak | Baseline | Current | Cap |",
                "| --- | ---: | ---: | ---: |",
            ]
        )
        baseline_peak_incidence = baseline_metrics.get("peak_incidence", {})
        for peak_id in sorted(peak_caps.keys()):
            lines.append(
                f"| {peak_id} | {baseline_peak_incidence.get(peak_id, 0)} | {current_peak_incidence.get(peak_id, 0)} | {peak_caps[peak_id]} |"
            )
        if any(baseline_peak_incidence.get(peak_id, 0) > peak_caps[peak_id] * 2 for peak_id in peak_caps):
            lines.extend(
                [
                    "",
                    "## Note",
                    "",
                    "- The baseline workbook `表号` semantics do not appear to match the CG duty identifier one-to-one, because the inferred baseline peak-incidence count is much larger than the XML caps in at least one peak.",
                ]
            )

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_solution(
    instance_dir: Path,
    solution_dir: Path,
    out_dir: Path,
    baseline_xlsx: Optional[Path],
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    travel_time_rows = _read_csv(instance_dir / "travel_times.csv")
    slot_mission_rows = _read_csv(instance_dir / "slot_missions.csv")
    option_rows = _read_csv(instance_dir / "options.csv")
    selected_trip_rows = _read_csv(solution_dir / "selected_trips.csv")
    selected_duty_rows = _read_csv(solution_dir / "selected_duties.csv")
    first_car_rows = _read_csv(instance_dir / "first_car_targets.csv")
    peaks = _load_peaks(instance_dir)
    summary = _load_summary(solution_dir)

    speed_headers, speed_sheet_rows = _build_speed_rows(travel_time_rows)
    legacy_like_rows = _build_legacy_like_rows(instance_dir, selected_trip_rows, slot_mission_rows)
    if legacy_like_rows is not None:
        (
            planned_headers,
            planned_csv_rows,
            planned_sheet_rows,
            mission_headers,
            mission_csv_rows,
            mission_sheet_rows,
        ) = legacy_like_rows
    else:
        planned_headers, planned_csv_rows, planned_sheet_rows = _build_planned_rows(selected_trip_rows)
        mission_headers, mission_csv_rows, mission_sheet_rows = _build_mission_rows(selected_trip_rows, slot_mission_rows)

    _write_csv(out_dir / "result_spd.csv", speed_headers, [{header: row[idx] for idx, header in enumerate(speed_headers)} for row in speed_sheet_rows])
    _write_csv(out_dir / "result_planned.csv", planned_headers, planned_csv_rows)
    _write_csv(out_dir / "result_mission.csv", mission_headers, mission_csv_rows)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    speed_sheet = workbook.create_sheet("运行时间")
    _write_sheet(speed_sheet, speed_headers, speed_sheet_rows)
    planned_sheet = workbook.create_sheet("计划线数据")
    _write_sheet(planned_sheet, planned_headers, planned_sheet_rows)
    mission_sheet = workbook.create_sheet("任务线数据")
    _write_sheet(mission_sheet, mission_headers, mission_sheet_rows)

    xlsx_path = out_dir / "result.xlsx"
    xls_path = out_dir / "result.xls"
    workbook.save(xlsx_path)
    shutil.copyfile(xlsx_path, xls_path)

    current_metrics = _build_current_metrics(selected_trip_rows, selected_duty_rows, option_rows, len(mission_csv_rows), peaks, summary, first_car_rows)
    current_metrics["planned_rows"] = len(planned_csv_rows)
    current_metrics["mission_rows"] = len(mission_csv_rows)
    current_metrics["vehicle_count"] = len({_parse_int(row["琛ㄥ彿"]) for row in planned_csv_rows})
    baseline_metrics = _load_baseline_metrics(baseline_xlsx, peaks) if baseline_xlsx is not None and baseline_xlsx.exists() else None
    turnback_overlaps = _check_turnback_overlaps(solution_dir)
    _write_comparison_report(out_dir / "comparison_report.md", current_metrics, baseline_metrics, turnback_overlaps)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Python-compatible Excel/CSV outputs for the C++ CG solver.")
    parser.add_argument("--instance-dir", required=True, help="Exported CG instance directory.")
    parser.add_argument("--solution-dir", required=True, help="Solved CG solution directory.")
    parser.add_argument("--out", required=True, help="Output directory for result.xls/result.xlsx.")
    parser.add_argument("--baseline", default=None, help="Optional baseline Python workbook for comparison.")
    args = parser.parse_args()

    export_solution(
        Path(args.instance_dir).resolve(),
        Path(args.solution_dir).resolve(),
        Path(args.out).resolve(),
        Path(args.baseline).resolve() if args.baseline else None,
    )


if __name__ == "__main__":
    main()
